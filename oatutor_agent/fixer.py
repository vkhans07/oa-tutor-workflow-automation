"""
OATutor Fixer
Consumes structured Issue objects from the validator and applies fixes.

Two passes:
  1. Structural fixes  — deterministic, always safe (ID renumbering, clearing stray fields, etc.)
  2. Content fixes     — Gemini-powered (missing answers, taxonomy, KC, answerType)

Usage (standalone):
    python fixer.py <target.xlsx> [--sheet SHEET_NAME]

Typical pipeline usage:
    from validator import validate
    from fixer     import fix

    issues = validate(ws, return_issues=True)
    fix(ws, issues, client=gemini_client)
    wb.save(path)
"""

import re
import string
import os
import sys
import argparse
from dataclasses import dataclass, field as dc_field
from typing import Any

import openpyxl
import google.genai as genai
from google.genai import types
from dotenv import load_dotenv


# ── Issue dataclass (shared with validator) ──────────────────────────────────

@dataclass
class Issue:
    row:     int
    code:    str          # see FIXABLE_CODES / CONTENT_CODES below
    context: dict = dc_field(default_factory=dict)
    fixable: bool = True  # structural fixes are auto; content fixes need Gemini


# Issue codes the fixer handles structurally
STRUCTURAL_CODES = {
    'BAD_HINT_ID',
    'BAD_SCAFFOLD_ID',
    'BAD_DEPENDENCY',
    'STRAY_HINTID_ON_STEP',
    'STRAY_DEP_ON_STEP',
    'MC_CHOICES_ON_NON_MC',
    'HINT_HAS_ANSWER',
    'WHITESPACE',
    'COLUMN_SHIFT',
}

# Issue codes that require Gemini to fix
CONTENT_CODES = {
    'MISSING_ANSWER',
    'MISSING_ANSWER_TYPE',
    'MISSING_TAXONOMY',
    'MISSING_KC',
}

# Issue codes that are flagged but not auto-fixable
SKIP_CODES = {
    'MISSING_OER',
    'MISSING_STEPS',
    'BAD_PROBLEM_NAME',
}


# ── Column map (1-based, matching validator.py) ──────────────────────────────

C = {
    'Problem Name': 1, 'Row Type': 2, 'Title': 3, 'Body Text': 4,
    'Answer': 5, 'answerType': 6, 'HintID': 7, 'Dependency': 8,
    'mcChoices': 9, 'Images': 10, 'Parent': 11, 'OER src': 12,
    'openstax KC': 13, 'KC': 14, 'Taxonomy': 15, 'License': 16,
    'Validator Check': 19,
}

EXPECTED_FIELDS = 16

HINT_ID_RE   = re.compile(r'^[hs]\d+$')
VALID_ROW_TYPES    = {'problem', 'step', 'hint', 'scaffold'}
VALID_ANSWER_TYPES = {'numeric', 'algebraic', 'algebra', 'mc', 'string'}

BLOOM_LEVELS = [
    'Remember', 'Understand', 'Apply', 'Analyze', 'Evaluate', 'Create'
]


# ── Cell helpers ─────────────────────────────────────────────────────────────

def gv(ws, row: int, col_name: str) -> str:
    v = ws.cell(row=row, column=C[col_name]).value
    return str(v).strip() if v is not None else ''

def sv(ws, row: int, col_name: str, value):
    ws.cell(row=row, column=C[col_name]).value = value

def is_blank(ws, row: int) -> bool:
    return all(
        ws.cell(row=row, column=c).value in (None, '')
        for c in range(1, EXPECTED_FIELDS + 1)
    )


# ── Structural fixers ────────────────────────────────────────────────────────

def _fix_bad_hint_id(ws, issue: Issue) -> str:
    expected = issue.context.get('expected', '')
    sv(ws, issue.row, 'HintID', expected)
    return f"HintID → {expected!r}"

def _fix_bad_scaffold_id(ws, issue: Issue) -> str:
    expected = issue.context.get('expected', '')
    sv(ws, issue.row, 'HintID', expected)
    return f"ScaffoldID → {expected!r}"

def _fix_bad_dependency(ws, issue: Issue) -> str:
    expected = issue.context.get('expected', '')
    sv(ws, issue.row, 'Dependency', expected or None)
    return f"Dependency → {expected!r}"

def _fix_stray_hintid(ws, issue: Issue) -> str:
    sv(ws, issue.row, 'HintID', None)
    return "cleared HintID on step row"

def _fix_stray_dep(ws, issue: Issue) -> str:
    sv(ws, issue.row, 'Dependency', None)
    return "cleared Dependency on step row"

def _fix_mc_choices_on_non_mc(ws, issue: Issue) -> str:
    sv(ws, issue.row, 'mcChoices', None)
    return "cleared mcChoices (answerType is not mc)"

def _fix_hint_has_answer(ws, issue: Issue) -> str:
    sv(ws, issue.row, 'Answer', None)
    sv(ws, issue.row, 'answerType', None)
    return "cleared Answer and answerType on hint row"

def _fix_whitespace(ws, issue: Issue) -> str:
    col = issue.context.get('col')
    if col:
        v = ws.cell(row=issue.row, column=C[col]).value
        if v:
            ws.cell(row=issue.row, column=C[col]).value = str(v).strip()
    return f"stripped whitespace in {col}"

def _fix_column_shift(ws, issue: Issue) -> str:
    """
    Reinsert a missing delimiter in a row that Gemini under-delimited.
    Uses the same scoring approach as generate_hints.py.
    """
    row_values = [
        ws.cell(row=issue.row, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    # Reconstruct as semicolon-delimited string and run repair
    raw_line = ';'.join('' if v is None else str(v) for v in row_values[:EXPECTED_FIELDS])
    repaired = _repair_single_row(raw_line)
    if repaired == raw_line:
        return "column shift: no repair found"
    fields = repaired.split(';')
    for i, val in enumerate(fields[:EXPECTED_FIELDS], start=1):
        ws.cell(row=issue.row, column=i).value = val or None
    return "column shift repaired"


def _repair_single_row(line: str) -> str:
    fields = line.split(';')
    if len(fields) == EXPECTED_FIELDS:
        return line

    def score(f):
        if len(f) != EXPECTED_FIELDS:
            return -1
        s = 0
        if f[1].strip().lower() in VALID_ROW_TYPES:            s += 3
        if f[5].strip().lower() in VALID_ANSWER_TYPES | {''}:  s += 2
        if HINT_ID_RE.match(f[6].strip()) or not f[6].strip(): s += 2
        if HINT_ID_RE.match(f[7].strip()) or not f[7].strip(): s += 1
        return s

    best, best_score = line, -1
    for pos in range(len(fields) + 1):
        candidate = fields[:pos] + [''] + fields[pos:]
        s = score(candidate)
        if s > best_score:
            best_score = s
            best = ';'.join(candidate)
    return best


STRUCTURAL_DISPATCH = {
    'BAD_HINT_ID':            _fix_bad_hint_id,
    'BAD_SCAFFOLD_ID':        _fix_bad_scaffold_id,
    'BAD_DEPENDENCY':         _fix_bad_dependency,
    'STRAY_HINTID_ON_STEP':   _fix_stray_hintid,
    'STRAY_DEP_ON_STEP':      _fix_stray_dep,
    'MC_CHOICES_ON_NON_MC':   _fix_mc_choices_on_non_mc,
    'HINT_HAS_ANSWER':        _fix_hint_has_answer,
    'WHITESPACE':             _fix_whitespace,
    'COLUMN_SHIFT':           _fix_column_shift,
}


# ── Gemini content fix prompts ───────────────────────────────────────────────

_MISSING_ANSWER_PROMPT = string.Template("""
You are filling in a missing answer for a single OATutor scaffold row.

CONTEXT
Problem: $problem_title
Step question: $step_title
Hint leading to this scaffold: $hint_body
Scaffold question: $scaffold_body

TASK
Return exactly two lines and nothing else:
Line 1: the answer (concise, LaTeX preferred for math expressions)
Line 2: the answerType — must be exactly one of: numeric, algebraic, mc

If answerType is mc, append a third line with 4 pipe-delimited choices where the correct answer is one of them.
Do not include labels, explanations, markdown, or blank lines.
""")

_MISSING_ANSWER_TYPE_PROMPT = string.Template("""
You are identifying the correct answerType for an OATutor row.

CONTEXT
Row type: $row_type
Answer value: $answer

TASK
Return exactly one word — the answerType. Must be one of: numeric, algebraic, mc
Do not include labels, explanations, or blank lines.
""")

_MISSING_TAXONOMY_PROMPT = string.Template("""
You are classifying an OATutor problem using Bloom's Taxonomy.

CONTEXT
Problem title: $problem_title
Step questions: $step_titles

TASK
Return exactly one Bloom's Taxonomy level from this list:
Remember, Understand, Apply, Analyze, Evaluate, Create

Choose the level that best describes the cognitive demand of the step questions.
Return only the single word — no explanation, no punctuation.
""")

_MISSING_KC_PROMPT = string.Template("""
You are assigning a knowledge component (KC) tag to an OATutor problem.

CONTEXT
Problem title: $problem_title
Step questions: $step_titles
OER source: $oer_src

TASK
Return a short KC tag (2–5 words, lowercase, hyphen-separated) that names the
mathematical concept being practised. Examples: "partial-fraction-decomposition",
"vector-magnitude", "cosine-difference-formula".

Return only the tag — no explanation, no punctuation, no quotes.
""")


# ── Gemini content fixers ────────────────────────────────────────────────────

def _get_problem_context(ws, row: int) -> dict:
    """Walk backwards to find the problem row, then collect step titles."""
    prob_row = row
    while prob_row > 2:
        if gv(ws, prob_row, 'Row Type').lower() == 'problem':
            break
        prob_row -= 1

    step_titles = []
    r = prob_row + 1
    while r <= ws.max_row and not is_blank(ws, r):
        if gv(ws, r, 'Row Type').lower() == 'step':
            step_titles.append(gv(ws, r, 'Title'))
        if gv(ws, r, 'Problem Name') != gv(ws, prob_row, 'Problem Name'):
            break
        r += 1

    return {
        'problem_title': gv(ws, prob_row, 'Title'),
        'oer_src':       gv(ws, prob_row, 'OER src'),
        'step_titles':   '\n'.join(f'- {t}' for t in step_titles),
    }


def _get_hint_context(ws, row: int) -> dict:
    """Find the step and preceding hint for a scaffold row."""
    step_title = hint_body = ''
    r = row - 1
    while r >= 2:
        rt = gv(ws, r, 'Row Type').lower()
        if rt == 'hint' and not hint_body:
            hint_body = gv(ws, r, 'Body Text') or gv(ws, r, 'Title')
        if rt == 'step':
            step_title = gv(ws, r, 'Title')
            break
        r -= 1
    return {'step_title': step_title, 'hint_body': hint_body}


def _call_gemini(client, prompt: str) -> str:
    resp = client.models.generate_content(
        model='gemini-2.5-flash',
        contents=prompt,
        config=types.GenerateContentConfig(temperature=0.0),
    )
    return resp.text.strip()


def _content_fix_missing_answer(ws, issue: Issue, client) -> str:
    ctx  = _get_problem_context(ws, issue.row)
    hint = _get_hint_context(ws, issue.row)
    prompt = _MISSING_ANSWER_PROMPT.substitute(
        problem_title  = ctx['problem_title'],
        step_title     = hint['step_title'],
        hint_body      = hint['hint_body'],
        scaffold_body  = gv(ws, issue.row, 'Body Text') or gv(ws, issue.row, 'Title'),
    )
    raw = _call_gemini(client, prompt)
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    if not lines:
        return "Gemini returned empty response — skipped"

    answer      = lines[0] if len(lines) > 0 else ''
    answer_type = lines[1] if len(lines) > 1 else ''
    mc_choices  = lines[2] if len(lines) > 2 else ''

    if answer_type not in VALID_ANSWER_TYPES:
        return f"Gemini returned invalid answerType {answer_type!r} — skipped"

    sv(ws, issue.row, 'Answer',     answer)
    sv(ws, issue.row, 'answerType', answer_type)
    if answer_type == 'mc' and mc_choices:
        sv(ws, issue.row, 'mcChoices', mc_choices)

    return f"filled Answer={answer!r} answerType={answer_type!r}"


def _content_fix_missing_answer_type(ws, issue: Issue, client) -> str:
    prompt = _MISSING_ANSWER_TYPE_PROMPT.substitute(
        row_type = gv(ws, issue.row, 'Row Type'),
        answer   = gv(ws, issue.row, 'Answer'),
    )
    raw = _call_gemini(client, prompt).strip().lower()
    if raw not in VALID_ANSWER_TYPES:
        return f"Gemini returned invalid answerType {raw!r} — skipped"
    sv(ws, issue.row, 'answerType', raw)
    return f"filled answerType={raw!r}"


def _content_fix_missing_taxonomy(ws, issue: Issue, client) -> str:
    ctx = _get_problem_context(ws, issue.row)
    prompt = _MISSING_TAXONOMY_PROMPT.substitute(
        problem_title = ctx['problem_title'],
        step_titles   = ctx['step_titles'],
    )
    raw = _call_gemini(client, prompt).strip()
    # Normalise capitalisation
    match = next((b for b in BLOOM_LEVELS if b.lower() == raw.lower()), None)
    if not match:
        return f"Gemini returned unrecognised Bloom's level {raw!r} — skipped"
    sv(ws, issue.row, 'Taxonomy', match)
    return f"filled Taxonomy={match!r}"


def _content_fix_missing_kc(ws, issue: Issue, client) -> str:
    ctx = _get_problem_context(ws, issue.row)
    prompt = _MISSING_KC_PROMPT.substitute(
        problem_title = ctx['problem_title'],
        step_titles   = ctx['step_titles'],
        oer_src       = ctx['oer_src'],
    )
    raw = _call_gemini(client, prompt).strip().lower()
    # Basic sanity: should be short and not contain spaces (use hyphens)
    if len(raw) > 60 or '\n' in raw:
        return f"Gemini returned unexpected KC value {raw!r} — skipped"
    sv(ws, issue.row, 'KC', raw)
    return f"filled KC={raw!r}"


CONTENT_DISPATCH = {
    'MISSING_ANSWER':      _content_fix_missing_answer,
    'MISSING_ANSWER_TYPE': _content_fix_missing_answer_type,
    'MISSING_TAXONOMY':    _content_fix_missing_taxonomy,
    'MISSING_KC':          _content_fix_missing_kc,
}


# ── Main fix entry point ─────────────────────────────────────────────────────

def fix(
    ws,
    issues: list[Issue],
    client=None,
    skip_content: bool = False,
) -> list[str]:
    """
    Apply all fixes to a live openpyxl worksheet.

    Args:
        ws:            openpyxl worksheet (modified in place)
        issues:        list of Issue objects from the validator
        client:        Gemini client for content fixes (optional — skipped if None)
        skip_content:  if True, skip content fixes even if client is provided

    Returns:
        list of changelog strings describing every change made
    """
    changelog = []
    skipped   = []

    structural = [i for i in issues if i.code in STRUCTURAL_CODES]
    content    = [i for i in issues if i.code in CONTENT_CODES]
    unfixable  = [i for i in issues if i.code in SKIP_CODES]

    # ── Pass 1: structural fixes ─────────────────────────────────────────────
    for issue in structural:
        fn = STRUCTURAL_DISPATCH.get(issue.code)
        if fn is None:
            skipped.append(f"Row {issue.row}: no handler for {issue.code!r}")
            continue
        try:
            msg = fn(ws, issue)
            changelog.append(f"[FIXED]   Row {issue.row} ({issue.code}): {msg}")
        except Exception as e:
            skipped.append(f"Row {issue.row} ({issue.code}): error — {e}")

    # ── Pass 2: content fixes (Gemini) ───────────────────────────────────────
    run_content = content and not skip_content and client is not None
    if content and client is None and not skip_content:
        print("  No Gemini client provided — skipping content fixes.")
        skipped.extend(
            f"Row {i.row} ({i.code}): needs Gemini (no client)" for i in content
        )
    elif run_content:
        print(f"\n  Running Gemini content fixes ({len(content)} issue(s))...")
        for issue in content:
            fn = CONTENT_DISPATCH.get(issue.code)
            if fn is None:
                skipped.append(f"Row {issue.row}: no handler for {issue.code!r}")
                continue
            try:
                msg = fn(ws, issue, client)
                changelog.append(f"[GEMINI]  Row {issue.row} ({issue.code}): {msg}")
            except Exception as e:
                skipped.append(f"Row {issue.row} ({issue.code}): Gemini error — {e}")

    # ── Unfixable ────────────────────────────────────────────────────────────
    for issue in unfixable:
        skipped.append(f"Row {issue.row} ({issue.code}): manual review required")

    # ── Summary ──────────────────────────────────────────────────────────────
    print(f"\n  Fixed: {len(changelog)}   Skipped: {len(skipped)}")
    if skipped:
        print("  Needs manual review:")
        for s in skipped:
            print(f"    ⚠ {s}")

    return changelog


# ── Standalone entry point ───────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='OATutor fixer')
    parser.add_argument('target',        help='Path to the .xlsx file to fix')
    parser.add_argument('--sheet',       default=None, help='Sheet name (default: active)')
    parser.add_argument('--no-content',  action='store_true', help='Skip Gemini content fixes')
    args = parser.parse_args()

    # Import here to avoid circular dependency when used as a module
    from validator import validate

    wb = openpyxl.load_workbook(args.target)
    ws = wb[args.sheet] if args.sheet else wb.active

    print("Running validator...")
    issues = validate(ws, return_issues=True)
    print(f"  {len(issues)} issue(s) found.\n")

    client = None
    if not args.no_content:
        load_dotenv()
        key = os.getenv('GEMINI_API_KEY')
        if key:
            client = genai.Client(api_key=key)
        else:
            print("  GEMINI_API_KEY not set — content fixes will be skipped.")

    changelog = fix(ws, issues, client=client, skip_content=args.no_content)

    wb.save(args.target)
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    print(f"\nSaved. {len(changelog)} fix(es) applied.")
    for line in changelog:
        print(f"  {line}")