"""
OATutor Spreadsheet Validator
Usage: python validator.py <target.xlsx> [--gold gold.xlsx] [--fix] [--sheet SHEET_NAME]

Checks every problem group against the OATutor spec and writes results to the
'Validator Check' column. With --fix, also corrects auto-fixable issues in place.
"""

import sys
import re
import argparse
import datetime
import openpyxl
from dataclasses import dataclass, field as dc_field


@dataclass
class Issue:
    row:     int
    code:    str
    context: dict = dc_field(default_factory=dict)
    fixable: bool = True

# Column mapping for easy reference
C = {
    'Problem Name': 1, 'Row Type': 2, 'Title': 3, 'Body Text': 4,
    'Answer': 5, 'answerType': 6, 'HintID': 7, 'Dependency': 8,
    'mcChoices': 9, 'Images': 10, 'Parent': 11, 'OER src': 12,
    'openstax KC': 13, 'KC': 14, 'Taxonomy': 15, 'License': 16,
    'Validator Check': 19,
}

VALID_ROW_TYPES    = {'problem', 'step', 'hint', 'scaffold'}
VALID_ANSWER_TYPES = {'numeric', 'algebra', 'mc'}

LATEX_SIGNALS  = re.compile(r'\\[a-zA-Z]+|\\frac|\\sqrt|\^{|_{|\$')
PYTHON_SIGNALS = re.compile(r'\*\*|(?<![=!<>])==(?!=)')

"""
Looks up a cell value with cv(row, col_name) and writes with sv(ws, row_num, col_name, value).
"""
def cv(row, col_name):
    idx = C.get(col_name)
    if idx is None or idx > len(row):
        return ''
    v = row[idx - 1]
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return f'__DATETIME_{v.month}/{v.day}__'
    return str(v).strip()

def sv(ws, row_num, col_name, value):
    ws.cell(row=row_num, column=C[col_name]).value = value

def is_blank_row(row):
    return all((v is None or str(v).strip() == '') for v in row)

# at the moment assuming only one format per sheet,
# could later extend to handle formatting conflicts
def detect_format(text):
    if LATEX_SIGNALS.search(text):
        return 'latex'
    if PYTHON_SIGNALS.search(text):
        return 'python'
    return 'plain'

# ─── Main validator ─────────────────────────────────────────────────────────

def validate(target_path, auto_fix=False, sheet_name=None, return_issues=False):
    wb = openpyxl.load_workbook(target_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if not ws:
        if sheet_name:
            raise ValueError(f"Sheet {sheet_name!r} not found in {target_path}")
        else:
            raise ValueError(f"No active sheet found in {target_path}")

    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        print("Sheet is empty.")
        return

    # ── Group rows into problem blocks ───────────────────────────────────────
    blocks = []
    current_block_start = None
    current_block = []

    for i, row in enumerate(raw_rows[1:], start=2):
        if is_blank_row(row):
            if current_block_start is not None:
                blocks.append((current_block_start, list(current_block)))
                current_block_start = None
                current_block = []
            continue
        rt = str(row[C['Row Type'] - 1]).strip().lower() if row[C['Row Type'] - 1] else ''
        if rt == 'problem':
            if current_block_start is not None:
                blocks.append((current_block_start, list(current_block)))
            current_block_start = i
            current_block = [i]
        elif current_block_start is not None:
            current_block.append(i)

    if current_block_start is not None:
        blocks.append((current_block_start, list(current_block)))

    # ── Detect dominant sheet format for consistency check ───────────────────
    all_text = []
    for row in raw_rows[1:]:
        for col in [C['Title']-1, C['Body Text']-1, C['Answer']-1]:
            v = row[col]
            if v and isinstance(v, str):
                all_text.append(v)
    sheet_format = detect_format(' '.join(all_text))

    issues    = {}   # prob_row -> [msg]
    row_issues = {}  # row_num  -> [msg]
    fixes     = []

    def add(prob_row, msg):
        issues.setdefault(prob_row, []).append(msg)

    def addr(row_num, msg):
        row_issues.setdefault(row_num, []).append(msg)

    for prob_row, block_rows in blocks:
        prow = raw_rows[prob_row - 1]
        prob_name = cv(prow, 'Problem Name')

        # ── Check 1: every problem must have at least one step ───────────
        step_rows = [
            r for r in block_rows
            if str(raw_rows[r-1][C['Row Type']-1] or '').strip().lower() == 'step'
        ]
        if not step_rows:
            add(prob_row, 'Problem has no steps')

        # ── Check 2: OER src, KC, Taxonomy populated ─────────────────────
        if not cv(prow, 'OER src'):
            add(prob_row, 'Missing OER src')
        if not cv(prow, 'KC'):
            add(prob_row, 'Missing KC')
        if not cv(prow, 'Taxonomy'):
            add(prob_row, 'Missing Taxonomy')

        # ── Check 3: Problem Name ends in a number (correct numbering) ───
        if not re.search(r'\d+$', prob_name):
            add(prob_row, f'Problem Name {prob_name!r} does not end in a number')

        # ── Check 4: problem row must not carry step-level fields ─────────
        if cv(prow, 'HintID') or cv(prow, 'answerType'):
            add(prob_row, 'Problem row has HintID or answerType (should be empty)')

        # ── Check 5: hint/scaffold IDs are correct per step sub-block ────
        sub_blocks = []
        current_sub = None
        for rn in block_rows:
            rt = str(raw_rows[rn-1][C['Row Type']-1] or '').strip().lower()
            if rt == 'step':
                if current_sub:
                    sub_blocks.append(current_sub)
                current_sub = [rn]
            elif rt in ('hint', 'scaffold') and current_sub is not None:
                current_sub.append(rn)
        if current_sub:
            sub_blocks.append(current_sub)

        for sub in sub_blocks:
            h_count = 0
            s_count = 0
            for rn in sub[1:]:
                row   = raw_rows[rn - 1]
                rt    = str(row[C['Row Type']-1] or '').strip().lower()
                hid   = cv(row, 'HintID')
                dep   = cv(row, 'Dependency')
                ans   = cv(row, 'Answer')
                atype = cv(row, 'answerType')
                mc    = cv(row, 'mcChoices')

                if rt == 'hint':
                    h_count += 1
                    expected_hid = f'h{h_count}'

                    if hid != expected_hid:
                        addr(rn, f'HintID {hid!r} → expected {expected_hid!r}')
                        if auto_fix:
                            sv(ws, rn, 'HintID', expected_hid)
                            fixes.append(f'Row {rn}: HintID {hid!r} → {expected_hid!r}')

                    if dep and dep.startswith('h'):
                        try:
                            if int(dep[1:]) >= h_count:
                                addr(rn, f'Hint Dependency {dep!r} refers to a future hint')
                                if auto_fix:
                                    expected_dep = f'h{h_count - 1}' if h_count > 1 else None
                                    sv(ws, rn, 'Dependency', expected_dep or None)
                                    fixes.append(f'Row {rn}: Dependency {dep!r} → {expected_dep!r}')
                        except ValueError:
                            addr(rn, f'Hint Dependency {dep!r} is not a valid hint ID')

                    if ans:
                        addr(rn, f'Hint has non-empty Answer {ans!r} (should be empty)')

                elif rt == 'scaffold':
                    s_count += 1
                    expected_sid = f's{s_count}'

                    if hid != expected_sid:
                        addr(rn, f'ScaffoldID {hid!r} → expected {expected_sid!r}')
                        if auto_fix:
                            sv(ws, rn, 'HintID', expected_sid)
                            fixes.append(f'Row {rn}: ScaffoldID {hid!r} → {expected_sid!r}')

                    if not ans:
                        addr(rn, 'Scaffold missing Answer')

                    if atype not in VALID_ANSWER_TYPES:
                        addr(rn, f'answerType {atype!r} not valid (numeric/algebra/mc)')

                    if atype == 'mc':
                        if not mc:
                            addr(rn, 'answerType=mc but mcChoices is empty')
                        else:
                            choices = [c.strip() for c in mc.split('|') if c.strip()]
                            if not (2 <= len(choices) <= 4):
                                addr(rn, f'mcChoices has {len(choices)} options (expected 2–4)')
                            if ans and ans not in choices:
                                addr(rn, f'Answer {ans!r} not in mcChoices')
                    else:
                        if mc:
                            addr(rn, f'mcChoices set but answerType={atype!r}')

                    # Algebra scaffold answers must not have commas
                    if atype in ('algebra') and ans and ',' in ans:
                        step_title = cv(raw_rows[sub[0]-1], 'Title')
                        if 'a,b' not in step_title.lower() and 'as a list' not in step_title.lower():
                            addr(rn, f'Algebra scaffold answer contains comma: {ans!r}')

        # ── Check 6: step-level field checks ─────────────────────────────
        for rn in step_rows:
            row   = raw_rows[rn - 1]
            ans   = cv(row, 'Answer')
            atype = cv(row, 'answerType')
            hid   = cv(row, 'HintID')
            dep   = cv(row, 'Dependency')
            title = cv(row, 'Title')

            if hid:
                addr(rn, f'Step row has HintID {hid!r} (should be empty)')
                if auto_fix:
                    sv(ws, rn, 'HintID', None)
                    fixes.append(f'Row {rn}: cleared HintID on step')
            if dep:
                addr(rn, f'Step row has Dependency {dep!r} (should be empty)')
                if auto_fix:
                    sv(ws, rn, 'Dependency', None)
                    fixes.append(f'Row {rn}: cleared Dependency on step')

            if atype and atype not in VALID_ANSWER_TYPES:
                addr(rn, f'answerType {atype!r} not valid')

            # Algebra step answers must not have commas
            if atype in ('algebra', 'algebraic') and ans and ',' in ans:
                addr(rn, f'Algebra step answer contains comma: {ans!r}')

            # Format consistency: flag if this row uses a different format from the sheet
            fmt = detect_format(title + ' ' + ans)
            if fmt != 'plain' and sheet_format != 'plain' and fmt != sheet_format:
                addr(rn, f'Formatting inconsistency: uses {fmt!r} but sheet convention is {sheet_format!r}')

    # ── Write results to Validator Check column ──────────────────────────────
    for i in range(2, len(raw_rows) + 1):
        ws.cell(row=i, column=C['Validator Check']).value = None

    for prob_row, msgs in issues.items():
        ws.cell(row=prob_row, column=C['Validator Check']).value = '\n'.join(msgs)

    for rn, msgs in row_issues.items():
        existing = ws.cell(row=rn, column=C['Validator Check']).value or ''
        combined = '\n'.join([existing] + msgs) if existing else '\n'.join(msgs)
        ws.cell(row=rn, column=C['Validator Check']).value = combined

    wb.save(target_path)

    # ── Print summary ────────────────────────────────────────────────────────
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    total_issues = sum(len(v) for v in issues.values()) + sum(len(v) for v in row_issues.values())
    print(f"\nValidation complete — {len(blocks)} problems, {total_issues} issue(s)")
    print(f"Results written to 'Validator Check' column in {target_path}\n")

    if total_issues:
        print("Issues by problem:")
        for prob_row, block_rows in blocks:
            prob_name = cv(raw_rows[prob_row - 1], 'Problem Name')
            prob_msgs = issues.get(prob_row, [])
            sub_msgs  = [(rn, row_issues[rn]) for rn in block_rows if rn in row_issues]
            if prob_msgs or sub_msgs:
                print(f"\n  [{prob_name}] (row {prob_row})")
                for m in prob_msgs:
                    print(f"    ✗ {m}")
                for rn, msgs in sub_msgs:
                    for m in msgs:
                        print(f"    ✗ Row {rn}: {m}")
    else:
        print("  No issues found.")

    if auto_fix and fixes:
        print(f"\nAuto-fixed {len(fixes)} issue(s):")
        for f in fixes:
            print(f"  → {f}")

    if return_issues:
        return _build_issue_list(issues, row_issues)


def _build_issue_list(issues, row_issues):
    structured = []
    unfixable  = {'MISSING_OER', 'MISSING_STEPS', 'BAD_PROBLEM_NAME'}
    for pr, msgs in issues.items():
        for msg in msgs:
            code = _msg_to_code(msg)
            structured.append(Issue(row=pr, code=code, context={'msg': msg},
                                    fixable=code not in unfixable))
    for rn, msgs in row_issues.items():
        for msg in msgs:
            code = _msg_to_code(msg)
            structured.append(Issue(row=rn, code=code, context={'msg': msg},
                                    fixable=code not in unfixable))
    return structured


def _msg_to_code(msg):
    m = msg.lower()
    if 'hintid' in m and 'expected' in m:              return 'BAD_HINT_ID'
    if 'scaffoldid' in m and 'expected' in m:          return 'BAD_SCAFFOLD_ID'
    if 'dependency' in m and ('future' in m or 'expected' in m): return 'BAD_DEPENDENCY'
    if 'step row has hintid' in m:                     return 'STRAY_HINTID_ON_STEP'
    if 'step row has dependency' in m:                 return 'STRAY_DEP_ON_STEP'
    if 'mcchoices set but' in m:                       return 'MC_CHOICES_ON_NON_MC'
    if 'hint has non-empty answer' in m:               return 'HINT_HAS_ANSWER'
    if 'scaffold missing answer' in m:                 return 'MISSING_ANSWER'
    if 'answertype' in m and 'not valid' in m:         return 'MISSING_ANSWER_TYPE'
    if 'missing taxonomy' in m:                        return 'MISSING_TAXONOMY'
    if 'missing kc' in m:                              return 'MISSING_KC'
    if 'missing oer' in m:                             return 'MISSING_OER'
    if 'no steps' in m:                                return 'MISSING_STEPS'
    if 'does not end in a number' in m:                return 'BAD_PROBLEM_NAME'
    return 'UNKNOWN'


# ─── Entry point ────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='OATutor spreadsheet validator')
    parser.add_argument('target',          help='Path to the .xlsx file to validate')
    parser.add_argument('--gold',          default='gold.xlsx', help='Gold standard reference (default: gold.xlsx)')
    parser.add_argument('--fix',           action='store_true',  help='Auto-fix hint/scaffold IDs and step HintID/Dependency fields')
    parser.add_argument('--sheet',         default=None,         help='Sheet name to validate (default: active sheet)')
    args = parser.parse_args()
    validate(args.target, args.gold, auto_fix=args.fix, sheet_name=args.sheet)